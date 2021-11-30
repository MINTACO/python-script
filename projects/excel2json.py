#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import xlrd
import os
import json
from utils.os_utils import get_all_file_path


def read_xls(entry):
    # 实例化excel
    book = xlrd.open_workbook(entry)
    # 获取sheets，返回一个xlrd.sheet.Sheet()对象
    sheets = book.sheets()
    # 返回dict
    result = {}
    # 循环读取sheet数据
    for sheet in sheets:
        # 获取总行数
        rows = sheet.nrows
        # 循环读取每行数据
        sheet_data = []
        for i in range(1, rows):
            # print(sheet.row_values(i))
            # 数据组装dic+t格式
            data = dict(zip(sheet.row_values(0), sheet.row_values(i)))
            sheet_data.append(data)
        result[sheet.sheet_name] = sheet_data
    return result


def read_xlsx(entry):
    # 加载文件
    book = load_workbook(entry)
    # 获取sheet列表
    sheets = book.sheetnames
    # 返回dict
    result = {}
    # 循环读取sheet数据
    for sheet_name in sheets:
        # sheet name获取sheet：
        sheet = book[sheet_name]
        # 获取总行数
        rows = sheet.max_row
        # 获取总列数
        # cols = sheet.max_column
        # 获取表头
        head = [row for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True)][0]
        # 数据组装
        sheet_data = []
        for row in sheet.iter_rows(min_row=2, max_row=rows, values_only=True):
            data = dict(zip(head, row))
            # for key in list(data.keys()):
            #     if not data.get(key):
            #         del data[key]
            sheet_data.append(data)
        result[sheet_name] = sheet_data
    return result


def dict2json(dict_data, output, filename, is_format):
    output_dir = os.path.join(output, filename)

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for key in dict_data:
        output = os.path.join(output_dir, key + '.json')
        # w-覆盖写入 a-追加写入 r-只读
        with open(output, 'w', encoding='UTF-8') as fileobject:
            if is_format:
                fileobject.write(json.dumps(dict_data[key], indent=4, ensure_ascii=False))
            else:
                fileobject.write(json.dumps(dict_data[key], ensure_ascii=False))
    print('Excel: ' + filename + ' To Json Done!')


def is_excel(file_path):
    # 把路径分割成dirname和basename，返回一个元组
    basename = os.path.split(file_path)[1]
    # 分割路径，返回路径名和文件扩展名的元组
    (filename, ext) = os.path.splitext(basename)
    if ext == '':
        return False
    if ext != '.xlsx' and ext != '.xls':
        return False
    # 忽略临时文件
    if filename[0] == '~' or filename[0] == '!':
        return False
    return filename, ext


def excel2json(entry: str, output: str = './output', is_format: bool = True):
    """
    :param entry:
    :param output:
    :param is_format:
    :return:
    """
    # 获取entry所有文件路径
    file_path_list = []
    # 输入文件
    if os.path.isfile(entry):
        file_path_list.append(entry)
    elif os.path.isdir(entry):
        # 递归获取文件名-文件路径dict
        file_path_list = get_all_file_path(entry)

    for file_path in file_path_list:
        # 校验文件
        if not is_excel(file_path):
            continue
        # 获取文件名，后缀名
        (filename, ext) = is_excel(file_path)
        excel_dict = {}
        if ext == '.xlsx':
            excel_dict = read_xlsx(file_path)
        elif ext == '.xls':
            excel_dict = read_xls(file_path)
        # 输出文件到json
        dict2json(excel_dict, output, filename, is_format)

    print('All Done!')
